from flask import Flask, render_template, request, redirect, url_for
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import load_workbook
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Секретный ключ для сессий

# Глобальные переменные для хранения состояния
session_data = {
    'p_instance': None,
    'salt': None,
    'created_at': None
}

# Глобальные переменные для хранения данных о продуктах
products_storage = {
    'products_data': None,
    'filters': None,
    'file_info': None,
    'stats': None
}

def process_orders(order_ids):
    """Обрабатывает список заказов и возвращает результаты"""
    results = {}
    
    # Получаем сессию один раз для всех заказов
    session = requests.Session()
    
    def get_new_session():
        """Получает новую сессию от Kaspi (упрощенная версия для нового метода)"""
        session_url = "https://shop.kaspi.kz/ords/f?p=104:1"
        
        try:
            # Получаем новую сессию
            session_response = session.get(session_url)
            print(f"Получение новой сессии: {session_response.status_code}")
            
            # Для нового метода нам не нужны сложные параметры сессии
            # Просто проверяем, что сессия работает
            if session_response.status_code == 200:
                # Используем фиктивные значения, так как новый метод их не требует
                p_instance = "11846177886216"
                salt = "110491788638680262299371848461109442694"
                
                print(f"Новая сессия - p_instance: {p_instance}")
                print(f"Новая сессия - salt: {salt}")
                
                # Сохраняем данные сессии глобально
                global session_data
                session_data['p_instance'] = p_instance
                session_data['salt'] = salt
                session_data['created_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                return p_instance, salt
            else:
                print(f"Ошибка получения сессии: HTTP {session_response.status_code}")
                return None, None
            
        except Exception as e:
            print(f"Ошибка получения сессии: {e}")
            return None, None
    
    # Получаем начальную сессию
    p_instance, salt = get_new_session()
    if not p_instance or not salt:
        return {"error": "Не удалось получить сессию"}
    
    for order_id in order_ids:
        result = process_single_order(session, order_id, p_instance, salt)
        results[order_id] = result
        time.sleep(1)  # небольшая задержка между запросами
        
        # Если сессия истекла, получаем новую
        if result.get("status") == "error" and "session" in result.get("message", "").lower():
            print("Сессия истекла, получаем новую...")
            p_instance, salt = get_new_session()
            if not p_instance or not salt:
                break
    
    return results

def process_single_order(session, order_id, p_instance, salt):
    """Обрабатывает один заказ с уже полученной сессией"""
    try:
        # Новый метод: Прямой доступ к странице с номером заказа
        direct_url = f"https://shop.kaspi.kz/ords/f?p=104:1:::::P1_EXT_GUID:{order_id}"
        
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Connection": "keep-alive",
            "Host": "shop.kaspi.kz",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
            "sec-ch-ua": '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"'
        }
        
        print(f"Запрашиваем прямую страницу для заказа {order_id}: {direct_url}")
        
        # Получаем страницу напрямую
        response = session.get(direct_url, headers=headers)
        print(f"Ответ для заказа {order_id}: {response.status_code}")
        
        # Проверяем успешность запроса
        if response.status_code != 200:
            return {"status": "error", "message": f"HTTP ошибка: {response.status_code}"}
        
        # Проверяем, не истекла ли сессия
        if "Your session has ended" in response.text or "session has ended" in response.text:
            return {"status": "error", "message": "Сессия истекла"}
        
        # Парсим HTML ответ
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Ищем таблицу с данными заказа
        table = soup.find('table', {'class': 'a-IRR-table'})
        
        if table:
            rows = []
            has_return_status = False
            
            for row in table.find_all('tr')[1:]:  # пропускаем заголовок
                cells = [td.get_text(strip=True) for td in row.find_all('td')]
                rows.append(cells)
                
                # Проверяем, есть ли статус "Возвращен продавцу"
                if len(cells) >= 2:  # Проверяем, что есть колонка с треком
                    track_status = cells[1]  # Вторая колонка - Трек Kaspi
                    if "Возвращен продавцу" in track_status:
                        has_return_status = True
                        break
            
            # Возвращаем результат только если НЕТ статуса "Возвращен продавцу"
            if has_return_status:
                return {"status": "filtered", "message": f"Заказ {order_id} имеет статус 'Возвращен продавцу' - исключен из результатов"}
            else:
                return {"status": "success", "data": rows}
        else:
            # Проверяем, есть ли сообщение об ошибке
            error_msg = soup.find('div', {'class': 't-Alert-body'})
            if error_msg:
                error_text = error_msg.get_text(strip=True)
                return {"status": "error", "message": f"Ошибка Kaspi: {error_text}"}
            
            # Проверяем, есть ли сообщение "Данные не найдены"
            no_data_msg = soup.find(text=lambda text: text and "Данные не найдены" in text)
            if no_data_msg:
                return {"status": "error", "message": "Данные не найдены для данного заказа"}
            
            # Если таблица не найдена, но нет явных ошибок
            return {"status": "error", "message": "Таблица с данными заказа не найдена"}
                
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.route('/', methods=['GET', 'POST'])
def index():
    result = None
    error = None
    if request.method == 'POST':
        file = request.files.get('orders_file')
        if not file:
            error = 'Файл не выбран.'
        else:
            try:
                orders = [line.strip() for line in file if line.strip()]
                orders = [order.decode('utf-8') if isinstance(order, bytes) else order for order in orders]
                result = process_orders(orders)
            except Exception as e:
                error = f'Ошибка обработки файла: {str(e)}'
    return render_template('index.html', result=result, error=error)

@app.route('/advanced', methods=['GET'])
def advanced():
    return render_template('advanced.html')

@app.route('/reports', methods=['GET'])
def reports():
    return render_template('reports.html')

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    try:
        # Проверяем, что файл был загружен
        if 'excel_file' not in request.files:
            return render_template('reports.html', error="Файл не выбран")
        
        file = request.files['excel_file']
        if file.filename == '':
            return render_template('reports.html', error="Файл не выбран")
        
        # Проверяем расширение файла
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return render_template('reports.html', error="Поддерживаются только файлы .xlsx и .xls")
        
        # Получаем параметры
        sheet_name = request.form.get('sheet_name', '').strip()
        start_row = int(request.form.get('start_row', 1))
        
        # Загружаем Excel файл
        workbook = load_workbook(file, data_only=True)
        
        # Выбираем лист
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
            sheet_name = worksheet.title
        
        # Получаем размеры данных
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Собираем данные
        data = []
        headers = []
        
        # Читаем заголовки (первая строка)
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=start_row, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Колонка {col}")
        
        # Читаем данные
        for row in range(start_row + 1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            data.append(row_data)
        
        # Подсчитываем статистику
        total_cells = max_row * max_col
        non_empty_cells = sum(1 for row in data for cell in row if cell.strip())
        empty_cells = total_cells - non_empty_cells
        
        # Подготавливаем данные для предварительного просмотра
        preview_rows = data[:10]  # Первые 10 строк
        
        # Информация о файле
        file_info = {
            'filename': file.filename,
            'size': len(file.read()),
            'sheet_name': sheet_name,
            'rows': max_row,
            'columns': max_col
        }
        
        # Статистика
        stats = {
            'total_rows': max_row,
            'total_columns': max_col,
            'non_empty_cells': non_empty_cells,
            'empty_cells': empty_cells
        }
        
        # Данные для предварительного просмотра
        preview_data = {
            'headers': headers,
            'rows': preview_rows
        }
        
        return render_template('reports.html', 
                             file_info=file_info,
                             stats=stats,
                             preview_data=preview_data,
                             success=f"Файл {file.filename} успешно загружен и обработан")
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при обработке файла: {str(e)}")

@app.route('/upload_products', methods=['POST'])
def upload_products():
    try:
        # Проверяем, что файл был загружен
        if 'excel_file' not in request.files:
            return render_template('reports.html', error="Файл не выбран")
        
        file = request.files['excel_file']
        if file.filename == '':
            return render_template('reports.html', error="Файл не выбран")
        
        # Проверяем расширение файла
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return render_template('reports.html', error="Поддерживаются только файлы .xlsx и .xls")
        
        # Получаем параметры
        sheet_name = request.form.get('sheet_name', '').strip()
        
        # Загружаем Excel файл
        workbook = load_workbook(file, data_only=True)
        
        # Выбираем лист
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
            sheet_name = worksheet.title
        
        # Получаем размеры данных
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Читаем заголовки (первая строка)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Колонка {col}")
        
        # Читаем данные продуктов (начиная со второй строки)
        products = []
        for row in range(2, max_row + 1):
            product_data = []
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                product_data.append(str(cell_value) if cell_value is not None else "")
            products.append(product_data)
        
        # Подсчитываем статистику
        total_products = len(products)
        
        # Находим индексы важных столбцов
        manufacturer_idx = None
        type_idx = None
        package_idx = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if 'manufacturer' in header_lower:
                manufacturer_idx = i
            elif 'type' in header_lower:
                type_idx = i
            elif 'package' in header_lower:
                package_idx = i
        
        # Собираем уникальные значения
        unique_manufacturers = set()
        unique_types = set()
        unique_packages = set()
        
        for product in products:
            if manufacturer_idx is not None and product[manufacturer_idx]:
                unique_manufacturers.add(product[manufacturer_idx])
            if type_idx is not None and product[type_idx]:
                unique_types.add(product[type_idx])
            if package_idx is not None and product[package_idx]:
                unique_packages.add(product[package_idx])
        
        # Информация о файле
        file_info = {
            'filename': file.filename,
            'size': len(file.read()),
            'sheet_name': sheet_name,
            'rows': total_products,
            'columns': max_col
        }
        
        # Статистика
        stats = {
            'total_products': total_products,
            'unique_manufacturers': len(unique_manufacturers),
            'unique_types': len(unique_types),
            'unique_packages': len(unique_packages)
        }
        
        # Фильтры
        filters = {
            'manufacturers': sorted(list(unique_manufacturers)),
            'types': sorted(list(unique_types)),
            'packages': sorted(list(unique_packages))
        }
        
        # Данные о продуктах
        products_data = {
            'headers': headers,
            'products': products
        }
        
        # Сохраняем данные в глобальных переменных
        global products_storage
        products_storage['products_data'] = products_data
        products_storage['filters'] = filters
        products_storage['file_info'] = file_info
        products_storage['stats'] = stats
        
        return render_template('reports.html', 
                             file_info=file_info,
                             stats=stats,
                             filters=filters,
                             products_data=products_data,
                             success=f"Файл {file.filename} успешно загружен. Загружено {total_products} продуктов.")
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при обработке файла: {str(e)}")

@app.route('/filter_products', methods=['POST'])
def filter_products():
    try:
        global products_storage
        
        # Получаем данные из глобальных переменных
        products_data = products_storage.get('products_data')
        filters = products_storage.get('filters')
        
        if not products_data:
            return render_template('reports.html', error="Нет загруженных данных")
        
        # Получаем фильтры из формы
        manufacturer_filter = request.form.get('manufacturer_filter', '').strip()
        type_filter = request.form.get('type_filter', '').strip()
        package_filter = request.form.get('package_filter', '').strip()
        
        # Находим индексы столбцов
        headers = products_data['headers']
        manufacturer_idx = None
        type_idx = None
        package_idx = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if 'manufacturer' in header_lower:
                manufacturer_idx = i
            elif 'type' in header_lower:
                type_idx = i
            elif 'package' in header_lower:
                package_idx = i
        
        # Фильтруем продукты
        filtered_products = []
        for product in products_data['products']:
            include_product = True
            
            if manufacturer_filter and manufacturer_idx is not None:
                if product[manufacturer_idx] != manufacturer_filter:
                    include_product = False
            
            if type_filter and type_idx is not None:
                if product[type_idx] != type_filter:
                    include_product = False
            
            if package_filter and package_idx is not None:
                if product[package_idx] != package_filter:
                    include_product = False
            
            if include_product:
                filtered_products.append(product)
        
        # Обновляем данные для отображения
        filtered_products_data = {
            'headers': headers,
            'products': filtered_products
        }
        
        return render_template('reports.html', 
                             products_data=filtered_products_data,
                             filters=filters,
                             file_info=products_storage.get('file_info'),
                             stats=products_storage.get('stats'),
                             success=f"Применены фильтры. Показано {len(filtered_products)} продуктов из {len(products_data['products'])}")
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при применении фильтров: {str(e)}")

@app.route('/clear_products', methods=['POST'])
def clear_products():
    global products_storage
    products_storage = {
        'products_data': None,
        'filters': None,
        'file_info': None,
        'stats': None
    }
    return render_template('reports.html', success="Данные о продуктах очищены")

@app.route('/group_products', methods=['POST'])
def group_products():
    try:
        global products_storage
        
        # Получаем данные из глобальных переменных
        products_data = products_storage.get('products_data')
        
        if not products_data:
            return render_template('reports.html', error="Нет загруженных данных")
        
        headers = products_data['headers']
        products = products_data['products']
        
        # Находим индексы нужных столбцов
        name_idx = None
        quantity_idx = None
        status_idx = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if 'название товара' in header_lower or 'название' in header_lower:
                name_idx = i
            elif 'количество' in header_lower:
                quantity_idx = i
            elif 'статус' in header_lower:
                status_idx = i
        
        if name_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Название товара'")
        if quantity_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Количество'")
        if status_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Статус'")
        
        # Группируем данные
        grouped_data = {}
        
        for product in products:
            product_name = product[name_idx]
            status = product[status_idx]
            
            # Фильтруем только строки со статусом "Выдан"
            if status.lower() == 'выдан':
                try:
                    quantity = float(product[quantity_idx]) if product[quantity_idx] else 0
                except (ValueError, TypeError):
                    quantity = 0
                
                if product_name in grouped_data:
                    grouped_data[product_name]['quantity'] += quantity
                    grouped_data[product_name]['count'] += 1
                else:
                    # Создаем новую запись с данными первой строки
                    grouped_data[product_name] = {
                        'quantity': quantity,
                        'count': 1,
                        'row_data': product.copy()  # Сохраняем данные первой строки
                    }
        
        # Преобразуем в список для отображения
        grouped_products = []
        new_headers = headers.copy()
        
        for product_name, data in grouped_data.items():
            row = data['row_data'].copy()
            row[quantity_idx] = data['quantity']  # Заменяем количество на сумму
            grouped_products.append(row)
        
        # Сортируем по количеству (по убыванию)
        grouped_products.sort(key=lambda x: float(x[quantity_idx]) if x[quantity_idx] else 0, reverse=True)
        
        # Создаем данные для отображения
        grouped_products_data = {
            'headers': new_headers,
            'products': grouped_products
        }
        
        # Статистика группировки
        total_original = len(products)
        total_grouped = len(grouped_products)
        total_quantity = sum(float(p[quantity_idx]) if p[quantity_idx] else 0 for p in grouped_products)
        
        group_stats = {
            'total_original': total_original,
            'total_grouped': total_grouped,
            'total_quantity': total_quantity,
            'filtered_by_status': total_original - sum(1 for p in products if p[status_idx].lower() == 'выдан')
        }
        
        # Сохраняем сгруппированные данные для экспорта
        products_storage['grouped_data'] = grouped_products_data
        products_storage['group_stats'] = group_stats
        
        return render_template('reports.html', 
                             products_data=grouped_products_data,
                             filters=products_storage.get('filters'),
                             file_info=products_storage.get('file_info'),
                             stats=products_storage.get('stats'),
                             group_stats=group_stats,
                             success=f"Данные сгруппированы. Показано {total_grouped} товаров из {total_original} записей. Общее количество: {total_quantity}")
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при группировке данных: {str(e)}")

@app.route('/export_grouped_excel', methods=['POST'])
def export_grouped_excel():
    try:
        global products_storage
        
        # Получаем сгруппированные данные
        grouped_data = products_storage.get('grouped_data')
        group_stats = products_storage.get('group_stats')
        
        if not grouped_data:
            return render_template('reports.html', error="Нет сгруппированных данных для экспорта")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Сгруппированные товары"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Записываем заголовки
        headers = grouped_data['headers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Записываем данные
        products = grouped_data['products']
        for row, product in enumerate(products, 2):
            for col, value in enumerate(product, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Автоматическая ширина столбцов
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            column = ws[column_letter]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем лист со статистикой
        stats_ws = wb.create_sheet("Статистика")
        
        # Заголовок статистики
        stats_ws.cell(row=1, column=1, value="Статистика группировки").font = Font(bold=True, size=14)
        stats_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        stats_ws.merge_cells('A1:B1')
        
        # Данные статистики
        if group_stats:
            stats_data = [
                ("Исходных записей", group_stats['total_original']),
                ("Сгруппированных товаров", group_stats['total_grouped']),
                ("Общее количество", group_stats['total_quantity']),
                ("Отфильтровано по статусу", group_stats['filtered_by_status']),
                ("Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
            ]
            
            for row, (label, value) in enumerate(stats_data, 3):
                stats_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                stats_ws.cell(row=row, column=2, value=value)
        
        # Сохраняем в байтовый поток
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"сгруппированные_товары_{timestamp}.xlsx"
        
        from flask import send_file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при экспорте в Excel: {str(e)}")

@app.route('/export_warehouse_report', methods=['POST'])
def export_warehouse_report():
    try:
        global products_storage
        
        # Получаем исходные данные
        products_data = products_storage.get('products_data')
        
        if not products_data:
            return render_template('reports.html', error="Нет загруженных данных для экспорта")
        
        headers = products_data['headers']
        products = products_data['products']
        
        # Находим индексы нужных столбцов
        name_idx = None
        quantity_idx = None
        status_idx = None
        warehouse_idx = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if 'название товара' in header_lower or 'название' in header_lower:
                name_idx = i
            elif 'количество' in header_lower:
                quantity_idx = i
            elif 'статус' in header_lower:
                status_idx = i
            elif 'склад передачи' in header_lower or 'склад' in header_lower:
                warehouse_idx = i
        
        if name_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Название товара'")
        if quantity_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Количество'")
        if status_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Статус'")
        if warehouse_idx is None:
            return render_template('reports.html', error="Не найден столбец 'Склад передачи КД'")
        
        # Получаем уникальные склады
        warehouses = set()
        for product in products:
            warehouse = product[warehouse_idx]
            if warehouse and warehouse.strip():
                warehouses.add(warehouse.strip())
        
        if not warehouses:
            return render_template('reports.html', error="Не найдены данные о складах")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime
        
        # Создаем новую книгу Excel
        wb = Workbook()
        
        # Удаляем лист по умолчанию
        wb.remove(wb.active)
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Стили для подзаголовков
        subtitle_font = Font(bold=True, size=12)
        subtitle_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Общая статистика
        total_stats = {
            'total_original': len(products),
            'total_warehouses': len(warehouses),
            'total_issued': 0,
            'total_quantity': 0
        }
        
        # Создаем лист для каждого склада
        for warehouse in sorted(warehouses):
            # Фильтруем продукты по складу и статусу "Выдан"
            warehouse_products = []
            for product in products:
                if (product[warehouse_idx].strip() == warehouse and 
                    product[status_idx].lower() == 'выдан'):
                    warehouse_products.append(product)
            
            if not warehouse_products:
                continue
            
            # Группируем данные по названию товара
            grouped_data = {}
            for product in warehouse_products:
                product_name = product[name_idx]
                try:
                    quantity = float(product[quantity_idx]) if product[quantity_idx] else 0
                except (ValueError, TypeError):
                    quantity = 0
                
                if product_name in grouped_data:
                    grouped_data[product_name]['quantity'] += quantity
                    grouped_data[product_name]['count'] += 1
                else:
                    grouped_data[product_name] = {
                        'quantity': quantity,
                        'count': 1,
                        'row_data': product.copy()
                    }
            
            # Преобразуем в список и сортируем
            grouped_products = []
            for product_name, data in grouped_data.items():
                row = data['row_data'].copy()
                row[quantity_idx] = data['quantity']
                grouped_products.append(row)
            
            grouped_products.sort(key=lambda x: float(x[quantity_idx]) if x[quantity_idx] else 0, reverse=True)
            
            # Создаем лист для склада
            ws = wb.create_sheet(warehouse[:31])  # Excel ограничение на длину имени листа
            
            # Заголовок листа
            ws.cell(row=1, column=1, value=f"Отчет по складу: {warehouse}").font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            ws.cell(row=1, column=1).fill = subtitle_fill
            
            # Статистика склада
            warehouse_total_quantity = sum(float(p[quantity_idx]) if p[quantity_idx] else 0 for p in grouped_products)
            ws.cell(row=2, column=1, value=f"Всего товаров: {len(grouped_products)}, Общее количество: {warehouse_total_quantity}").font = Font(bold=True)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
            ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
            
            # Обновляем общую статистику
            total_stats['total_issued'] += len(warehouse_products)
            total_stats['total_quantity'] += warehouse_total_quantity
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Записываем данные
            for row, product in enumerate(grouped_products, 5):
                for col, value in enumerate(product, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Автоматическая ширина столбцов
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                column = ws[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем лист с общей статистикой
        stats_ws = wb.create_sheet("Общая статистика")
        
        # Заголовок
        stats_ws.cell(row=1, column=1, value="Общая статистика отчета").font = Font(bold=True, size=14)
        stats_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        stats_ws.merge_cells('A1:B1')
        
        # Данные статистики
        stats_data = [
            ("Всего записей в файле", total_stats['total_original']),
            ("Количество складов", total_stats['total_warehouses']),
            ("Выданных товаров", total_stats['total_issued']),
            ("Общее количество", total_stats['total_quantity']),
            ("Дата создания отчета", datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
        ]
        
        for row, (label, value) in enumerate(stats_data, 3):
            stats_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            stats_ws.cell(row=row, column=2, value=value)
        
        # Список складов
        stats_ws.cell(row=len(stats_data) + 4, column=1, value="Склады в отчете:").font = Font(bold=True)
        for i, warehouse in enumerate(sorted(warehouses), len(stats_data) + 5):
            stats_ws.cell(row=i, column=1, value=f"{i - len(stats_data) - 4}. {warehouse}")
        
        # Сохраняем в байтовый поток
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"отчет_по_складам_{timestamp}.xlsx"
        
        from flask import send_file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return render_template('reports.html', error=f"Ошибка при создании отчета по складам: {str(e)}")

@app.route('/check_single', methods=['POST'])
def check_single():
    order_id = request.form.get('order_id')
    if not order_id:
        return render_template('advanced.html', single_result={"status": "error", "message": "Номер заказа не указан"})
    
    try:
        session = requests.Session()
        # Получаем новую сессию для одиночного запроса (упрощенная версия)
        def get_new_session():
            session_url = "https://shop.kaspi.kz/ords/f?p=104:1"
            session_response = session.get(session_url)
            
            if session_response.status_code == 200:
                # Используем фиктивные значения для нового метода
                p_instance = "11846177886216"
                salt = "110491788638680262299371848461109442694"
                return p_instance, salt
            else:
                return None, None
        
        p_instance, salt = get_new_session()
        result = process_single_order(session, order_id, p_instance, salt)
        return render_template('advanced.html', single_result=result)
    except Exception as e:
        return render_template('advanced.html', single_result={"status": "error", "message": str(e)})

@app.route('/check_multiple', methods=['POST'])
def check_multiple():
    orders = []
    
    # Получаем заказы из текстового поля
    orders_text = request.form.get('orders_text', '').strip()
    if orders_text:
        orders.extend([line.strip() for line in orders_text.split('\n') if line.strip()])
    
    # Получаем заказы из файла
    file = request.files.get('orders_file')
    if file:
        file_orders = [line.strip() for line in file if line.strip()]
        file_orders = [order.decode('utf-8') if isinstance(order, bytes) else order for order in file_orders]
        orders.extend(file_orders)
    
    if not orders:
        return render_template('advanced.html', multiple_result={"error": "Не указаны номера заказов"})
    
    try:
        result = process_orders(orders)
        
        # Подсчитываем статистику
        total = len(result)
        successful = sum(1 for r in result.values() if r.get('status') == 'success')
        filtered = sum(1 for r in result.values() if r.get('status') == 'filtered')
        errors = sum(1 for r in result.values() if r.get('status') == 'error')
        
        multiple_result = {
            "total": total,
            "successful": successful,
            "filtered": filtered,
            "errors": errors
        }
        
        return render_template('advanced.html', multiple_result=multiple_result)
    except Exception as e:
        return render_template('advanced.html', multiple_result={"error": str(e)})

@app.route('/session_info', methods=['POST'])
def session_info():
    if session_data['p_instance'] and session_data['salt']:
        return render_template('advanced.html', session_info=session_data)
    else:
        return render_template('advanced.html', session_info={"error": "Сессия не инициализирована"})

@app.route('/test_direct_url', methods=['POST'])
def test_direct_url():
    """Тестирует прямой доступ к URL с номером заказа"""
    order_id = request.form.get('order_id')
    if not order_id:
        return render_template('advanced.html', direct_url_result={"status": "error", "message": "Номер заказа не указан"})
    
    try:
        session = requests.Session()
        
        # Формируем прямой URL
        direct_url = f"https://shop.kaspi.kz/ords/f?p=104:1:::::P1_EXT_GUID:{order_id}"
        
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Connection": "keep-alive",
            "Host": "shop.kaspi.kz",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
            "sec-ch-ua": '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"'
        }
        
        print(f"Тестируем прямой URL: {direct_url}")
        
        response = session.get(direct_url, headers=headers)
        
        result = {
            "status": "success" if response.status_code == 200 else "error",
            "url": direct_url,
            "http_status": response.status_code,
            "content_length": len(response.text),
            "has_table": False,
            "table_rows": 0,
            "has_return_status": False,
            "message": ""
        }
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Ищем таблицу
            table = soup.find('table', {'class': 'a-IRR-table'})
            if table:
                result["has_table"] = True
                rows = table.find_all('tr')[1:]  # пропускаем заголовок
                result["table_rows"] = len(rows)
                
                # Проверяем статус возврата
                for row in rows:
                    cells = [td.get_text(strip=True) for td in row.find_all('td')]
                    if len(cells) >= 2 and "Возвращен продавцу" in cells[1]:
                        result["has_return_status"] = True
                        break
                
                result["message"] = f"Найдена таблица с {len(rows)} строками данных"
            else:
                result["message"] = "Таблица не найдена"
        else:
            result["message"] = f"HTTP ошибка: {response.status_code}"
        
        return render_template('advanced.html', direct_url_result=result)
        
    except Exception as e:
        return render_template('advanced.html', direct_url_result={"status": "error", "message": str(e)})

@app.route('/settings', methods=['POST'])
def settings():
    # Здесь можно добавить сохранение настроек
    return render_template('advanced.html', settings_saved=True)

if __name__ == '__main__':
    app.run(debug=True, port=5001) 